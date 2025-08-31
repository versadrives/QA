import eventlet
eventlet.monkey_patch()
from app import app

from flask import Flask, render_template, request, redirect, send_file, jsonify
from flask_socketio import SocketIO
import sqlite3
from datetime import datetime
import random
import os, sys
import logging
from rs485_reader import get_live_power_and_factor_and_rpm

def resource_path(rel):
    try:
        base = sys._MEIPASS  # created by PyInstaller when frozen
    except Exception:
        base = os.path.abspath(".")
    return os.path.join(base, rel)

import shutil

def ensure_writable_db():
    target_path = os.path.join(os.path.abspath("."), "scan_log.db")  # or user folder
    if not os.path.exists(target_path):
        shutil.copy(resource_path("scan_log.db"), target_path)
    return target_path

DB_FILE = ensure_writable_db()


app = Flask(
    __name__,
    template_folder=resource_path("templates"),
    static_folder=resource_path("static"),
)

socketio = SocketIO(app, cors_allowed_origins="*", async_mode='eventlet')


DB_FILE = resource_path("scan_log.db")


logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

def init_db():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    # Create scans table
    c.execute("""
    CREATE TABLE IF NOT EXISTS scans (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        daily_number INTEGER NOT NULL,
        qr_code TEXT NOT NULL,
        power REAL NOT NULL,
        rpm INTEGER NOT NULL,
        power_factor REAL NOT NULL,
        failure_code TEXT NOT NULL,
        status TEXT NOT NULL,
        timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
        result TEXT DEFAULT 'FP OK',
        voice_recognition TEXT DEFAULT 'NA'
    )
    """)
    
    # ✅ Create models table to store specs per model
    c.execute("""
    CREATE TABLE IF NOT EXISTS models (
        model_prefix TEXT PRIMARY KEY,
        power_min REAL,
        power_max REAL,
        pf_min REAL,
        rpm_min INTEGER,
        rpm_max INTEGER
    )
""")


    # ✅ Create settings table if not exists
    c.execute("""
    CREATE TABLE IF NOT EXISTS settings (
        key TEXT PRIMARY KEY,
        value TEXT
    )
    """)

    # ✅ Insert default voice recognition setting if not exists
    c.execute("""
    INSERT OR IGNORE INTO settings (key, value)
    VALUES ('default_voice_recognition', 'NA')
    """)

    conn.commit()
    conn.close()


def get_db():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn

def insert_scan(qr_code, power=None, rpm=None, power_factor=None, failure_code='NA', result=None):
    try:
        # Get current voice recognition setting for new scan
        voice_recognition = get_default_voice_recognition()

        with get_db() as conn:
            cur = conn.cursor()
            today = datetime.now().strftime('%Y-%m-%d')
            cur.execute("""
                SELECT COUNT(*) + 1 as number 
                FROM scans 
                WHERE date(timestamp) = ?
            """, (today,))
            daily_number = cur.fetchone()['number']

            # REQUIRE power, rpm, power_factor explicitly passed; else error out
            if power is None or rpm is None or power_factor is None:
                raise ValueError("power, rpm, and power_factor must be provided")

            # Extract model prefix
            model_prefix = qr_code.split('.')[0]

            # Fetch model limits from DB
            cur.execute("""
                SELECT power_min, power_max, pf_min, rpm_min, rpm_max
                FROM models
                WHERE model_prefix = ? COLLATE NOCASE
            """, (model_prefix,))
            model = cur.fetchone()

            # Default to FAIL if model not found
            status = 'FAIL'
            if model:
                power_min, power_max, pf_min, rpm_min, rpm_max = model
                if (
                    power_min <= power <= power_max and
                    pf_min <= power_factor and
                    rpm_min <= rpm <= rpm_max
                ):
                    status = 'PASS'
                    failure_code = 'NA'
                    result = 'FP OK'
                else:
                    if failure_code == 'NA':
                        failure_code = ''
                    result = failure_code if result is None else result
            else:
                failure_code = 'UNKNOWN MODEL'
                result = 'MODEL NOT FOUND'

            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            cur.execute("""
                INSERT INTO scans (
                    daily_number, qr_code, power, rpm, power_factor, 
                    failure_code, status, timestamp, result, voice_recognition
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                daily_number, qr_code, power, rpm, power_factor,
                failure_code, status, timestamp, result, voice_recognition
            ))
            conn.commit()

            return {
                'daily_number': daily_number,
                'qr_code': qr_code,
                'power': power,
                'rpm': rpm,
                'power_factor': power_factor,
                'failure_code': failure_code,
                'status': status,
                'timestamp': timestamp,
                'result': result,
                'voice_recognition': voice_recognition
            }
    except Exception as e:
        print(f"Error inserting scan: {str(e)}")
        return None
    
def get_stats():
    today = datetime.now().strftime('%Y-%m-%d')
    current_month = datetime.now().strftime('%Y-%m')

    with get_db() as conn:
        cur = conn.cursor()

        # First Passed
        cur.execute("SELECT COUNT(*) FROM scans WHERE date(timestamp)=? AND result='FP OK'", (today,))
        first_passed = cur.fetchone()[0]

        # Second Passed
        cur.execute("SELECT COUNT(*) FROM scans WHERE date(timestamp)=? AND result='SP OK'", (today,))
        second_passed = cur.fetchone()[0]

        # Total Passed
        total_passed = first_passed + second_passed

        # Rework
        cur.execute("SELECT COUNT(*) FROM scans WHERE date(timestamp)=? AND result='RW'", (today,))
        rework = cur.fetchone()[0]

    return {
        "total_passed": total_passed,
        "first_passed": first_passed,
        "second_passed": second_passed,
        "rework": rework
    }


def get_scans(date=None):
    try:
        with get_db() as conn:
            cur = conn.cursor()
            query = """
                SELECT daily_number, qr_code, power, rpm, power_factor, failure_code, status, timestamp, result, voice_recognition
                FROM scans
            """
            params = []
            if date:
                query += " WHERE date(timestamp) = ?"
                params.append(date)
            query += " ORDER BY id DESC"
            cur.execute(query, params)
            return [dict(row) for row in cur.fetchall()]
    except Exception as e:
        print(f"Error getting scans: {str(e)}")
        return []

@app.route('/')
def index():
    try:
        date = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
        scans = get_scans(date)

        with get_db() as conn:
            cur = conn.cursor()

            # Calculate First Passed (today only) - result 'FP OK'
            cur.execute("""
                SELECT COUNT(*) as count FROM scans
                WHERE date(timestamp) = ? AND result = 'FP OK'
            """, (date,))
            first_passed = cur.fetchone()['count']

            # Calculate Second Passed (today only) - result 'SP OK'
            cur.execute("""
                SELECT COUNT(*) as count FROM scans
                WHERE date(timestamp) = ? AND result = 'SP OK'
            """, (date,))
            second_passed = cur.fetchone()['count']

            # Total Passed = First Passed + Second Passed
            total_passed = first_passed + second_passed

            # Calculate Rework (today only) - result 'RW'
            cur.execute("""
                SELECT COUNT(*) as count FROM scans
                WHERE date(timestamp) = ? AND result = 'RW'
            """, (date,))
            rework = cur.fetchone()['count']

            # Fetch models for settings modal
            cur.execute("SELECT * FROM models ORDER BY model_prefix")
            models = cur.fetchall()

        return render_template('index.html',
                               scans=scans,
                               selected_date=date,
                               total_passed=total_passed,
                               first_passed=first_passed,
                               rework=rework,
                               second_passed=second_passed,
                               models=models)
    except Exception as e:
        print(f"Error in index route: {str(e)}")
        return render_template('index.html',
                               scans=[],
                               selected_date=datetime.now().strftime('%Y-%m-%d'),
                               total_passed=0,
                               first_passed=0,
                               rework=0,
                               second_passed=0,
                               models=[])

@app.route('/scan', methods=['POST'])
def scan():
    qr_code = request.form.get('qr_code', '').strip()
    failure_code = request.form.get('failure_code', 'NA')
    if not qr_code:
        return jsonify({'error': 'QR code is required'}), 400
    
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT 1 FROM scans WHERE qr_code=? AND result='FP OK' LIMIT 1", (qr_code,))
        if cur.fetchone():
            return jsonify({
                'success': False,
                'duplicate_fp_ok': True,
                'message': 'Duplicate scan not allowed.'
            }), 200

    power, power_factor, rpm = get_live_power_and_factor_and_rpm()
    if power is None or power_factor is None or rpm is None:
        return jsonify({'error': 'Failed to read sensors data from RS485'}), 500

    scan_data = insert_scan(qr_code, power=power, rpm=rpm, power_factor=power_factor, failure_code=failure_code)
    if scan_data:
        stats = get_stats()  # 👈 get updated numbers
        socketio.emit('new_scan', {**scan_data, **stats})  # emit to all clients
        return jsonify({'success': True, 'data': scan_data, 'stats': stats})  # 👈 include stats in response
    else:
        return jsonify({'error': 'Failed to insert scan'}), 500

@app.route('/export')
def export():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    file_name = request.args.get('file_name', 'scan_report')
    
    if not all([start_date, end_date]):
        return jsonify({'error': 'Date range is required'}), 400
        
    try:
        with get_db() as conn:
            cur = conn.cursor()
            cur.execute("""
                SELECT daily_number, qr_code, power, rpm, power_factor, failure_code, result, voice_recognition
                FROM scans
                WHERE date(timestamp) BETWEEN ? AND ?
                ORDER BY timestamp ASC
            """, (start_date, end_date))
            
            scans = cur.fetchall()
            
            if not scans:
                return jsonify({'error': 'No data found'}), 404
                
            filename = os.path.join(os.path.expanduser("~/Documents"), f"{file_name}.xlsx")
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font

            wb = Workbook()
            ws = wb.active

            headers = [
                'SL.NO', 'QR Code', 'Power', 'RPM', 'Power Factor', 'Failure Code',
                'IS 302-1 A-3 Functional test',
                'IS 374  18.4 d Simple running  test',
                'IS 374 4.6 Enclosure',
                'IS 374 Cl. 4.3 Blades and Motor',
                'Result',
                'Voice Recognition'
            ]
            ws.append(headers)

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True)

            for scan in scans:
                row = [
                    scan['daily_number'],
                    scan['qr_code'],
                    scan['power'],
                    scan['rpm'],
                    scan['power_factor'],
                    scan['failure_code'],
                    'OK',
                    'OK',
                    'OK',
                    'OK',
                    scan['result'],
                    scan['voice_recognition']
                ]
                ws.append(row)
                for col_num in range(1, len(row) + 1):
                    ws.cell(row=ws.max_row, column=col_num).alignment = Alignment(horizontal='center', vertical='center')

            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            wb.save(filename)
            return send_file(filename, as_attachment=True)
    except Exception as e:
        print(f"Error in export: {str(e)}")
        return jsonify({'error': 'Export failed'}), 500
    
    

@app.route('/undo', methods=['POST'])
def undo():
    try:
        with get_db() as conn:
            cur = conn.cursor()
            cur.execute("SELECT id FROM scans ORDER BY id DESC LIMIT 1")
            last_scan = cur.fetchone()
            
            if last_scan:
                cur.execute("DELETE FROM scans WHERE id = ?", (last_scan['id'],))
                conn.commit()
                return jsonify({'success': True})
            else:
                return jsonify({'error': 'No scans to remove'}), 404
                
    except Exception as e:
        print(f"Error in undo: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/models', methods=['GET', 'POST'])
def manage_models():
    with get_db() as conn:
        cur = conn.cursor()
        
        if request.method == 'POST':
            action = request.form.get('action')
            prefix = request.form.get('model_prefix', '').strip()

            if action == 'add':
                power_min = float(request.form['power_min'])
                power_max = float(request.form['power_max'])
                pf_min = float(request.form['pf_min'])
                rpm_min = int(request.form['rpm_min'])
                rpm_max = int(request.form['rpm_max'])
                cur.execute("""
                    INSERT OR REPLACE INTO models 
                    (model_prefix, power_min, power_max, pf_min, rpm_min, rpm_max)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (prefix, power_min, power_max, pf_min, rpm_min, rpm_max))
                conn.commit()

            elif action == 'update':
                power_min = float(request.form['power_min'])
                power_max = float(request.form['power_max'])
                pf_min = float(request.form['pf_min'])
                rpm_min = int(request.form['rpm_min'])
                rpm_max = int(request.form['rpm_max'])
                cur.execute("""
                    UPDATE models 
                    SET power_min = ?, power_max = ?, pf_min = ?, rpm_min = ?, rpm_max = ?
                    WHERE model_prefix = ? COLLATE NOCASE
                """, (power_min, power_max, pf_min, rpm_min, rpm_max, prefix))
                conn.commit()

            elif action == 'delete':
                cur.execute("DELETE FROM models WHERE model_prefix = ? COLLATE NOCASE", (prefix,))
                conn.commit()

        # Fetch updated models
        cur.execute("SELECT * FROM models ORDER BY model_prefix")
        models = cur.fetchall()

        # Fetch scan data and stats for dashboard
        from datetime import datetime
        selected_date = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
        
        # Re-use your existing get_scans function or implement below
        scans = get_scans(selected_date)

        # Calculate stats
        current_month = selected_date[:7]
        cur.execute("""
            SELECT COUNT(*) as count 
            FROM scans WHERE strftime('%Y-%m', timestamp) = ?
        """, (current_month,))
        monthly_scans = cur.fetchone()['count']

        today_total = len([scan for scan in scans if scan['timestamp'].startswith(selected_date)])
        today_passed = len([scan for scan in scans if scan['timestamp'].startswith(selected_date) and scan['status'] == 'PASS'])
        today_failed = today_total - today_passed

    # Render the main dashboard template with full context data
    return render_template('index.html', 
                           scans=scans,
                           selected_date=selected_date,
                           models=models,
                           monthly_scans=monthly_scans,
                           today_total=today_total,
                           today_passed=today_passed,
                           today_failed=today_failed)



@app.route('/update_failure_code', methods=['POST'])
def update_failure_code():
    qr_code = request.form.get('qr_code', '').strip()
    failure_code = request.form.get('failure_code', '').strip()
    if not qr_code or not failure_code:
        return jsonify({'error': 'QR code and failure code are required'}), 400

    try:
        with get_db() as conn:
            cur = conn.cursor()
            # Update the most recent failed scan for this QR code with empty failure_code
            cur.execute("""
                UPDATE scans
                SET failure_code = ?
                WHERE qr_code = ? AND status = 'FAIL' AND (failure_code = '' OR failure_code IS NULL)
                ORDER BY id DESC
                LIMIT 1
            """, (failure_code, qr_code))
            conn.commit()
            if cur.rowcount == 0:
                return jsonify({'error': 'No matching failed scan found'}), 404
            return jsonify({'success': True})
    except Exception as e:
        print(f"Error updating failure code: {str(e)}")
        return jsonify({'error': 'Failed to update failure code'}), 500

def get_default_voice_recognition():
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT value FROM settings WHERE key = 'default_voice_recognition'")
        result = cur.fetchone()
        return result['value'] if result else 'NA'

@app.route('/voice_recognition', methods=['POST'])
def voice_recognition():
    option = request.form.get('option')
    if option not in ['OK', 'NA']:
        return jsonify({'error': 'Invalid option'}), 400
    try:
        with get_db() as conn:
            cur = conn.cursor()
            # Only update the settings table, not existing scans
            cur.execute("""
                INSERT OR REPLACE INTO settings (key, value)
                VALUES ('default_voice_recognition', ?)
            """, (option,))
            conn.commit()
        return jsonify({'success': True, 'selected': option})
    except Exception as e:
        print(f"Error updating voice recognition: {str(e)}")
        return jsonify({'error': str(e)}), 500
    
@app.route('/edit_last_scan', methods=['POST'])
def edit_last_scan():
    failure_code = request.form.get('failure_code', '').strip()
    result = request.form.get('result', '').strip()
    if not failure_code or not result:
        return jsonify({'error': 'Both failure code and result are required'}), 400
    try:
        with get_db() as conn:
            cur = conn.cursor()
            # Get last scan's ID
            cur.execute("SELECT id FROM scans ORDER BY id DESC LIMIT 1")
            last_scan = cur.fetchone()
            if not last_scan:
                return jsonify({'error': 'No scans found'}), 404
            scan_id = last_scan['id']
            # Update failure_code and result for last scan
            cur.execute("""
                UPDATE scans
                SET failure_code = ?, result = ?
                WHERE id = ?
            """, (failure_code, result, scan_id))
            conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    
@app.route('/last_scan')
def last_scan():
    try:
        with get_db() as conn:
            cur = conn.cursor()
            cur.execute("SELECT * FROM scans ORDER BY id DESC LIMIT 1")
            scan = cur.fetchone()
            if not scan:
                return jsonify({'error': 'No scans found'}), 404
            return jsonify(dict(scan))
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    

@app.route('/update_result', methods=['POST'])
def update_result():
    result = request.form.get('result', '').strip()
    if not result:
        return jsonify({'error': 'Result is required'}), 400
    try:
        with get_db() as conn:
            cur = conn.cursor()
            # Update the most recent scan
            cur.execute("""
                UPDATE scans
                SET result = ?
                WHERE id = (SELECT id FROM scans ORDER BY id DESC LIMIT 1)
            """, (result,))
            conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/update_failure_code_and_result', methods=['POST'])
def update_failure_code_and_result():
    try:
        failure_code = request.form.get('failure_code', '').strip()
        result = request.form.get('result', '').strip()

        logger.debug(f"Received failure_code: {failure_code}, result: {result}")

        if not failure_code:
            return jsonify({'error': 'Failure code is required'}), 400

        with get_db() as conn:
            cur = conn.cursor()
            # First, verify if the last scan exists and is a failure
            cur.execute("""
                SELECT id FROM scans 
                WHERE status = 'FAIL' 
                ORDER BY id DESC LIMIT 1
            """)
            last_scan = cur.fetchone()
            
            if not last_scan:
                return jsonify({'error': 'No failed scan found'}), 404

            # Update both failure code and result
            cur.execute("""
                UPDATE scans 
                SET failure_code = ?, result = ?
                WHERE id = ?
            """, (failure_code, result, last_scan['id']))
            
            conn.commit()

            # Verify the update
            if cur.rowcount == 0:
                return jsonify({'error': 'Update failed'}), 500

            return jsonify({'success': True})

    except Exception as e:
        print(f"Error in update_failure_code_and_result: {str(e)}")
        return jsonify({'error': str(e)}), 500
@app.route('/defaults')
def defaults():
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT value FROM settings WHERE key = 'default_voice_recognition'")
        result = cur.fetchone()
        voice_default = result['value'] if result else 'NA'
    return jsonify({'default_voice_recognition': voice_default})

@app.route('/clear_scans', methods=['POST'])
def clear_scans():
    try:
        with get_db() as conn:
            cur = conn.cursor()
            cur.execute("DELETE FROM scans")  # clears only scan logs
            conn.commit()
        return jsonify({'success': True, 'message': 'All scan logs cleared successfully.'})
    except Exception as e:
        print(f"Error clearing scans: {str(e)}")
        return jsonify({'error': 'Failed to clear scans'}), 500


if __name__ == '__main__':
    init_db()
    print(">>> Flask-SocketIO async_mode:", socketio.async_mode)  # debug print
    socketio.run(app, host='0.0.0.0', port=5000, debug=True)
